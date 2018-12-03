import openpyxl
from openpyxl.styles import Font
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet   
# from the active attribute. 
sheet = wb.active
  
# set the height of the row 
sheet.row_dimensions[1].height = 70
  
# set the width of the column 
sheet.column_dimensions['B'].width = 20

sheet.merge_cells('A2:D4')  
sheet.cell(row = 2, column = 1).value = 'Twelve cells join together.'
  
# merge cell C6 and D6 
sheet.merge_cells('F6:G6')  
sheet.cell(row = 6, column = 6).value = 'Two merge cells.'

#unmerge cells
#sheet.unmerge_cells('A2:D4')  
#sheet.unmerge_cells('C6:D6') 


sheet.cell(row = 8, column = 1).value = "Paul Burkholder"
  
# set the size of the cell to 24 
sheet.cell(row = 8, column = 1).font = Font(size = 24 ) 
  
sheet.cell(row = 9, column = 1).value = "Paul Burkholder"
  
# set the font style to italic 
sheet.cell(row = 9, column = 1).font = Font(size = 24, italic = True) 
  
sheet.cell(row = 10, column = 1).value = "Paul Burkholder"
  
# set the font style to bold 
sheet.cell(row = 10, column = 1).font = Font(size = 24, bold = True) 
  
sheet.cell(row = 11, column = 1).value = "Paul Burkholder"
  
# set the font name to 'Times New Roman' 
sheet.cell(row = 11, column = 1).font = Font(size = 24, name = 'Times New Roman')

wb.save("pyxlsx_Formatting.xlsx") 
