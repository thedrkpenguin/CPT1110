import openpyxl
  
def search_inventory():
    path = "inventory.xlsx"
  
    #Open the workbook
    wb_obj = openpyxl.load_workbook(path) 
  
    #Get the active sheet in the file.
    sheet_obj = wb_obj.active

    #Get object to search for
    search_object = int(input("Enter the ID number of the item: "))

    #Provide starting location
    cell_obj = sheet_obj.cell(row = 2, column = 1)

    #Get maximum number of rows
    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column

    for i in range(2,max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        id_number = cell_obj.value
        if id_number != search_object:
            print("Item not found!")
        else:
            print("Item found!")
            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row = 2, column = i)
                print(cell_obj.value, end = " ")
            break


if __name__ == "__main__":
    search_inventory()
    
